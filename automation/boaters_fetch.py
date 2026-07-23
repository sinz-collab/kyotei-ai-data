# -*- coding: utf-8 -*-
"""
BOATERS 全12R 統合取得 v3.0
ENTRYクリック取得 + DATA + MOTOR / 1-6・7-12分割 / 各競艇場フォルダ保存版

v2.3修正点:
- v2.2の「出走表」クリック取得は維持
- ENTRYページ遷移後に下方向へ段階スクロールして出走表本体テーブルを強制ロード
- 出走表判定を強化:
  全国勝率 / 当地勝率 / 平均ST / モーター / ボート / 節間成績 を確認
- ENTRY保存前に body.innerText を複数回チェックし、読み込み不足を減らす
- data/motorは従来通り取得

使い方:
py boaters_all_races_merge_v2_2.py --stadium ashiya --date 2026-05-24 --headed

保存先:
C:\\Users\\sinz1\\kyotei_ai\\競艇場のレース\\芦屋\\20260524\\
"""

import argparse
import asyncio
import json
import re
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
from zoneinfo import ZoneInfo
from html import unescape

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None


BASE_SAVE_ROOT = Path("work/races")

STADIUM_MAP = {
    "ashiya": "芦屋",
    "shimonoseki": "下関",
    "wakamatsu": "若松",
    "tokoname": "常滑",
    "omura": "大村",
    "karatsu": "唐津",
    "heiwajima": "平和島",
    "toda": "戸田",
    "edogawa": "江戸川",
    "fukuoka": "福岡",
    "芦屋": "芦屋",
    "下関": "下関",
    "若松": "若松",
    "常滑": "常滑",
    "大村": "大村",
    "唐津": "唐津",
    "平和島": "平和島",
    "戸田": "戸田",
    "江戸川": "江戸川",
    "福岡": "福岡",
}

STADIUM_SLUG_MAP = {
    "芦屋": "ashiya",
    "下関": "shimonoseki",
    "若松": "wakamatsu",
    "常滑": "tokoname",
    "大村": "omura",
    "唐津": "karatsu",
    "平和島": "heiwajima",
    "戸田": "toda",
    "江戸川": "edogawa",
    "福岡": "fukuoka",
}

DATA_CLICK_STEPS = [
    {"label": "AI3連対率・枠別勝率", "keywords": ["AI3連対率・枠別勝率", "AI3連対率", "枠別勝率"]},
    {"label": "枠別スタート順位", "keywords": ["枠別スタート順位"]},
    {"label": "決まり手率", "keywords": ["決まり手率"]},
    {"label": "前づけデータ", "keywords": ["前づけデータ"]},
    {"label": "先頭艇別連対率", "keywords": ["先頭艇別連対率"]},
    {"label": "直近10走", "keywords": ["直近10走"]},
    {"label": "直近6ヶ月", "keywords": ["直近6ヶ月"]},
    {"label": "当地", "keywords": ["当地"]},
    {"label": "今期", "keywords": ["今期"]},
    {"label": "一般戦", "keywords": ["一般戦"]},
]

DANGER_WORDS = [
    "ログイン", "会員登録", "トップ", "データベース", "AI予想実績",
    "レース結果・払戻", "お問い合わせ", "利用規約", "プライバシー",
    "キャンペーン", "ランキング"
]

ENTRY_KEYWORDS = [
    "出走表",
    "枠",
    "レーサー",
    "選手",
    "級別",
    "全国勝率",
    "当地勝率",
    "全国2連率",
    "当地2連率",
    "平均ST",
    "F",
    "L",
    "モーター",
    "ボート",
    "節間成績",
    "早見",
    "年齢",
    "支部",
    "体重",
]

DATA_KEYWORDS = [
    "AI3連対率・枠別勝率",
    "AI3連対率",
    "枠別勝率",
    "枠別スタート順位",
    "決まり手率",
    "前づけデータ",
    "先頭艇別連対率",
    "直近10走",
    "当地",
]

MOTOR_KEYWORDS = [
    "モーター情報",
    "モーター直近10走",
    "モーター",
    "2連対率",
    "3連対率",
    "展示タイム",
    "展示順位",
    "優出",
    "優勝",
    "出走回数",
]


def normalize_stadium(stadium: str) -> str:
    if stadium not in STADIUM_MAP:
        raise ValueError(f"未対応の競艇場です: {stadium}")
    return STADIUM_MAP[stadium]


def stadium_to_slug(stadium: str) -> str:
    jp = normalize_stadium(stadium)
    return STADIUM_SLUG_MAP[jp]


def today_jst() -> str:
    return datetime.now(ZoneInfo("Asia/Tokyo")).strftime("%Y-%m-%d")


def date_to_folder(date: str) -> str:
    return date.replace("-", "")


def build_url(slug: str, date: str, race_no: int, page_name: str) -> str:
    return f"https://boaters-boatrace.com/race/{slug}/{date}/{race_no}R/{page_name}"


async def get_body_text(page) -> str:
    try:
        return await page.evaluate("() => document.body ? document.body.innerText : ''")
    except Exception:
        return ""


async def save_current(page, race_dir: Path, base_name: str) -> Dict:
    html = ""
    text = ""
    try:
        html = await page.content()
    except Exception:
        pass
    try:
        text = await get_body_text(page)
    except Exception:
        pass

    html_path = race_dir / f"{base_name}.html"
    txt_path = race_dir / f"{base_name}.txt"
    html_path.write_text(html, encoding="utf-8", errors="ignore")
    txt_path.write_text(text, encoding="utf-8", errors="ignore")

    return {
        "html": str(html_path),
        "txt": str(txt_path),
        "text_len": len(text),
        "text": text,
    }


def keyword_hits(text: str, keywords: List[str]) -> List[str]:
    return [kw for kw in keywords if kw in text]


def analyze_entry_text(text: str) -> Dict:
    is_first_day = "初日" in text
    hits = keyword_hits(text, ENTRY_KEYWORDS)

    # 出走表としてかなり強い条件
    strong_words = ["全国勝率", "当地勝率", "平均ST", "モーター", "ボート"]
    strong_hit_count = sum(1 for w in strong_words if w in text)

    return {
        "text_len": len(text),
        "is_first_day": is_first_day,
        "setsukan_status": "初日のため未発生" if is_first_day else ("節間成績あり候補" if "節間" in text or "節間成績" in text else "要確認"),
        "hits": hits,
        "strong_hit_count": strong_hit_count,
        "has_entry_core": ("出走表" in text and strong_hit_count >= 2),
        "has_motor_boat_refs": ("モーター" in text and "ボート" in text),
    }


def analyze_data_text(text: str) -> Dict:
    hits = keyword_hits(text, DATA_KEYWORDS)
    return {
        "text_len": len(text),
        "hits": hits,
        "has_data_core": ("AI3連対率" in text and "決まり手率" in text),
    }


def analyze_motor_text(text: str) -> Dict:
    hits = keyword_hits(text, MOTOR_KEYWORDS)
    return {
        "text_len": len(text),
        "hits": hits,
        "has_motor_core": ("モーター情報" in text and "2連対率" in text and "モーター直近10走" in text),
    }


async def goto_page(page, url: str, wait_ms: int) -> Dict:
    result = {"url": url, "warning": None}
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    except PlaywrightTimeoutError as e:
        result["warning"] = f"timeout: {e}"
    except Exception as e:
        result["warning"] = f"error: {e}"

    await page.wait_for_timeout(wait_ms)
    return result


async def find_best_click_target(page, keywords):
    script = r"""
    (args) => {
      const keywords = args.keywords;
      const dangerWords = args.dangerWords;

      const selectors = [
        'button',
        'a',
        '[role="tab"]',
        '[role="button"]',
        'li',
        'div[class*="tab"]',
        'div[class*="Tab"]',
        'span'
      ];

      const nodes = Array.from(document.querySelectorAll(selectors.join(',')));
      const out = [];

      function visible(el) {
        const r = el.getBoundingClientRect();
        const st = window.getComputedStyle(el);
        return r.width > 5 && r.height > 5 && st.visibility !== 'hidden' && st.display !== 'none';
      }

      for (const el of nodes) {
        if (!visible(el)) continue;

        const text = (el.innerText || el.textContent || '').trim().replace(/\s+/g, ' ');
        const aria = (el.getAttribute('aria-label') || '').trim();
        const href = (el.getAttribute('href') || '').trim();
        const title = (el.getAttribute('title') || '').trim();
        const joined = [text, aria, href, title].join(' ').trim();

        if (!joined) continue;

        let danger = false;
        for (const d of dangerWords) {
          if (joined.includes(d)) danger = true;
        }
        if (danger) continue;

        let score = 0;
        for (const kw of keywords) {
          if (text === kw) score += 100;
          else if (text.includes(kw)) score += 30;
          else if (joined.includes(kw)) score += 10;
        }
        if (score <= 0) continue;

        if (text.length > 180) score -= 25;
        if (text.length > 400) score -= 80;

        const r = el.getBoundingClientRect();
        out.push({
          text,
          aria,
          href,
          title,
          score,
          x: r.x + r.width / 2,
          y: r.y + r.height / 2,
          width: r.width,
          height: r.height,
          tag: el.tagName,
          role: el.getAttribute('role') || '',
          className: String(el.className || '')
        });
      }

      out.sort((a,b) => b.score - a.score || a.text.length - b.text.length || a.y - b.y);
      return out[0] || null;
    }
    """
    try:
        return await page.evaluate(script, {"keywords": keywords, "dangerWords": DANGER_WORDS})
    except Exception as e:
        return {"error": repr(e)}


async def click_target_and_wait(page, target: Dict, wait_ms: int) -> Dict:
    before = page.url
    res = {"before_url": before, "clicked": False, "target": target}

    try:
        async with page.expect_navigation(wait_until="domcontentloaded", timeout=8000) as nav_info:
            await page.mouse.click(float(target["x"]), float(target["y"]))
            res["clicked"] = True
        try:
            nav = await nav_info.value
            res["navigation_url"] = nav.url if nav else None
        except Exception:
            pass
    except Exception:
        # SPA遷移/同ページ切替/ナビなしの場合
        try:
            await page.mouse.click(float(target["x"]), float(target["y"]))
            res["clicked"] = True
        except Exception as e:
            res["click_error"] = repr(e)

    await page.wait_for_timeout(wait_ms)
    res["after_url"] = page.url
    return res



async def force_load_entry_body(page, total_wait_ms: int = 9000) -> Dict:
    """
    ENTRYページ内の出走表本体を読み込ませるため、段階スクロール＋待機を行う。
    SPA/遅延描画対策。
    """
    checkpoints = []
    waited = 0

    async def snapshot(label: str):
        text = await get_body_text(page)
        analysis = analyze_entry_text(text)
        checkpoints.append({
            "label": label,
            "text_len": len(text),
            "strong_hit_count": analysis.get("strong_hit_count"),
            "hits": analysis.get("hits", []),
            "has_entry_core": analysis.get("has_entry_core"),
        })
        return analysis

    analysis = await snapshot("initial_entry")
    if analysis.get("has_entry_core") and analysis.get("strong_hit_count", 0) >= 4:
        return {"loaded": True, "checkpoints": checkpoints}

    scroll_steps = [0, 700, 1400, 2200, 3200, 4400, 5600, 7000, -7000]
    for i, y in enumerate(scroll_steps):
        try:
            if y == 0:
                await page.mouse.wheel(0, -10000)
            else:
                await page.mouse.wheel(0, y)
            await page.wait_for_timeout(900)
            waited += 900
        except Exception:
            pass

        analysis = await snapshot(f"scroll_{i}_{y}")
        if analysis.get("has_entry_core") and analysis.get("strong_hit_count", 0) >= 4:
            return {"loaded": True, "checkpoints": checkpoints}

        if waited >= total_wait_ms:
            break

    try:
        await page.wait_for_load_state("networkidle", timeout=5000)
    except Exception:
        pass

    analysis = await snapshot("final_after_networkidle")
    return {"loaded": bool(analysis.get("has_entry_core")), "checkpoints": checkpoints}



async def capture_entry_by_click(page, slug: str, date: str, race_no: int, race_dir: Path, wait_ms: int) -> Dict:
    """
    /dataを開いて「出走表」タブをクリックし、実際の出走表ページを保存する。
    """
    data_url = build_url(slug, date, race_no, "data")
    print(f"[ENTRY {race_no}R] open data then click 出走表: {data_url}")

    nav = await goto_page(page, data_url, wait_ms)

    # 上部に戻してタブを探す
    try:
        await page.mouse.wheel(0, -5000)
        await page.wait_for_timeout(300)
    except Exception:
        pass

    target = await find_best_click_target(page, ["出走表"])

    # 見つからない場合は少しスクロールしながら探す
    if not target:
        for _ in range(3):
            try:
                await page.mouse.wheel(0, 800)
                await page.wait_for_timeout(300)
            except Exception:
                pass
            target = await find_best_click_target(page, ["出走表"])
            if target:
                break

    click_res = {"target": target, "clicked": False}
    if target and "error" not in target:
        click_res = await click_target_and_wait(page, target, wait_ms)
    else:
        click_res["error"] = "出走表タブが見つかりません"

    # 遷移後に出走表本体・節間/下部情報もロード
    entry_load = await force_load_entry_body(page)

    saved = await save_current(page, race_dir, f"race_{race_no:02d}_entry")
    analysis = analyze_entry_text(saved["text"])

    # もしクリック後もdataのままで出走表コアが弱い場合、hrefから直接URLを開く試行
    direct_attempt = None
    href = target.get("href") if isinstance(target, dict) else ""
    if (not analysis["has_entry_core"]) and href:
        direct_url = href
        if href.startswith("/"):
            direct_url = "https://boaters-boatrace.com" + href
        print(f"[ENTRY {race_no}R] direct href retry: {direct_url}")
        await goto_page(page, direct_url, wait_ms)
        href_entry_load = await force_load_entry_body(page)
        saved2 = await save_current(page, race_dir, f"race_{race_no:02d}_entry_href_retry")
        analysis2 = analyze_entry_text(saved2["text"])
        direct_attempt = {
            "url": direct_url,
            "analysis": analysis2,
            "entry_load": href_entry_load,
            "saved": {k: v for k, v in saved2.items() if k != "text"},
        }

        if analysis2.get("strong_hit_count", 0) > analysis.get("strong_hit_count", 0):
            # 正式ファイルへ上書き
            (race_dir / f"race_{race_no:02d}_entry.html").write_text(
                (race_dir / f"race_{race_no:02d}_entry_href_retry.html").read_text(encoding="utf-8", errors="ignore"),
                encoding="utf-8", errors="ignore"
            )
            (race_dir / f"race_{race_no:02d}_entry.txt").write_text(
                (race_dir / f"race_{race_no:02d}_entry_href_retry.txt").read_text(encoding="utf-8", errors="ignore"),
                encoding="utf-8", errors="ignore"
            )
            saved = saved2
            analysis = analysis2

    return {
        "race_no": race_no,
        "ok": True,
        "start_url": data_url,
        "current_url": page.url,
        "nav": nav,
        "click": click_res,
        "entry_load": entry_load,
        "direct_attempt": direct_attempt,
        "analysis": analysis,
        "saved": {k: v for k, v in saved.items() if k != "text"},
    }


async def click_data_tabs(page, click_wait_ms: int) -> List[Dict]:
    results = []

    for step in DATA_CLICK_STEPS:
        label = step["label"]
        try:
            await page.mouse.wheel(0, -5000)
            await page.wait_for_timeout(200)
        except Exception:
            pass

        target = await find_best_click_target(page, step["keywords"])

        if not target:
            for _ in range(4):
                try:
                    await page.mouse.wheel(0, 1200)
                    await page.wait_for_timeout(300)
                except Exception:
                    pass
                target = await find_best_click_target(page, step["keywords"])
                if target:
                    break

        res = {"label": label, "target": target, "clicked": False}

        if target and "error" not in target:
            before = page.url
            try:
                await page.mouse.click(float(target["x"]), float(target["y"]))
                res["clicked"] = True
                await page.wait_for_timeout(click_wait_ms)

                after = page.url
                res["before_url"] = before
                res["after_url"] = after

                base_before = before.split("#")[0]
                base_after = after.split("#")[0]
                res["unsafe_navigation"] = base_after != base_before

                if res["unsafe_navigation"]:
                    try:
                        await page.go_back(wait_until="domcontentloaded", timeout=15_000)
                        await page.wait_for_timeout(800)
                    except Exception as e:
                        res["go_back_error"] = repr(e)
            except Exception as e:
                res["error"] = repr(e)

        results.append(res)

    return results


async def capture_data_page(page, slug: str, date: str, race_no: int, race_dir: Path, wait_ms: int, click_wait_ms: int) -> Dict:
    url = build_url(slug, date, race_no, "data")
    print(f"[DATA {race_no}R] {url}")

    nav = await goto_page(page, url, wait_ms)
    clicks = await click_data_tabs(page, click_wait_ms)

    try:
        await page.mouse.wheel(0, 2000)
        await page.wait_for_timeout(600)
        await page.mouse.wheel(0, -2000)
        await page.wait_for_timeout(300)
    except Exception:
        pass

    saved = await save_current(page, race_dir, f"race_{race_no:02d}_data")
    analysis = analyze_data_text(saved["text"])

    return {
        "race_no": race_no,
        "ok": True,
        "url": url,
        "current_url": page.url,
        "nav": nav,
        "clicks": clicks,
        "analysis": analysis,
        "saved": {k: v for k, v in saved.items() if k != "text"},
    }


async def capture_motor_page(page, slug: str, date: str, race_no: int, race_dir: Path, wait_ms: int) -> Dict:
    url = build_url(slug, date, race_no, "motor")
    print(f"[MOTOR {race_no}R] {url}")

    nav = await goto_page(page, url, wait_ms)

    try:
        await page.mouse.wheel(0, 2400)
        await page.wait_for_timeout(800)
        await page.mouse.wheel(0, -2400)
        await page.wait_for_timeout(300)
    except Exception:
        pass

    saved = await save_current(page, race_dir, f"race_{race_no:02d}_motor")
    analysis = analyze_motor_text(saved["text"])

    return {
        "race_no": race_no,
        "ok": True,
        "url": url,
        "current_url": page.url,
        "nav": nav,
        "analysis": analysis,
        "saved": {k: v for k, v in saved.items() if k != "text"},
    }



def strip_html_to_clean_text(html: str) -> str:
    """
    AI予想投入用の軽量テキスト化。
    元HTMLは残すため、これは削除型ではなく「別出力の整形版」。
    """
    if not html:
        return ""

    text = html

    # 明確なノイズのみ除去
    text = re.sub(r"(?is)<script\b[^>]*>.*?</script>", "\n", text)
    text = re.sub(r"(?is)<style\b[^>]*>.*?</style>", "\n", text)
    text = re.sub(r"(?is)<svg\b[^>]*>.*?</svg>", "\n", text)
    text = re.sub(r"(?is)<iframe\b[^>]*>.*?</iframe>", "\n", text)
    text = re.sub(r"(?is)<noscript\b[^>]*>.*?</noscript>", "\n", text)
    text = re.sub(r"(?is)<ins\b[^>]*>.*?</ins>", "\n", text)

    # タグ境界を改行にする
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</(p|div|li|tr|td|th|h1|h2|h3|h4|section|article|button|a|span)>", "\n", text)
    text = re.sub(r"(?is)<[^>]+>", "\n", text)

    text = unescape(text)

    # Next.js内部JSON系も読めるよう最低限改行
    replacements = [
        ('","', '"\n"'),
        ('},{', '}\n{'),
        (',"__typename"', '\n"__typename"'),
        (',"racerName"', '\n"racerName"'),
        (',"racerRank"', '\n"racerRank"'),
        (',"motorN"', '\n"motorN"'),
        (',"boatN"', '\n"boatN"'),
        (',"kimarite"', '\n"kimarite"'),
        (',"aggType"', '\n"aggType"'),
        (',"tenjiTime"', '\n"tenjiTime"'),
        (',"tenjiRank"', '\n"tenjiRank"'),
        (',"sinnyu"', '\n"sinnyu"'),
        (',"result"', '\n"result"'),
    ]
    for a, b in replacements:
        text = text.replace(a, b)

    noise_patterns = [
        "googletagmanager", "googleads", "doubleclick", "Advertisement",
        "chakra-colors", "chakra-space", "data-google", "recaptcha",
        "_next/static", "favicons", "manifest.json", "origin-trial",
        "YouTubeチャンネル", "お問い合わせ・ご意見", "利用規約",
        "プライバシーポリシー", "特定商取引法", "運営会社",
        "font-family", "rgba(", "color-scheme", "data-emotion"
    ]

    keep_when_long = [
        "racerName", "motorN", "boatN", "kimarite", "aiProba",
        "CrawledRaceRacer", "MotorRecentResult", "WakuAggregation",
        "DecisionRateAggregation", "resultIs1AvgWithWaku", "result3renAvgWithWaku"
    ]

    lines = []
    blank = 0
    for raw in text.splitlines():
        line = re.sub(r"\s+", " ", raw.strip())
        if not line:
            blank += 1
            if blank <= 1:
                lines.append("")
            continue
        blank = 0

        if any(p in line for p in noise_patterns):
            continue

        if len(line) > 700 and not any(k in line for k in keep_when_long):
            continue

        lines.append(line)

    # 連続重複を圧縮
    compact = []
    prev = None
    rep = 0
    for line in lines:
        if line == prev:
            rep += 1
            if rep <= 1:
                compact.append(line)
        else:
            rep = 0
            compact.append(line)
        prev = line

    return "\n".join(compact).strip()


def make_clean_range(out_dir: Path, race_dir: Path, start_race: int, end_race: int, suffix: str, log: Dict):
    """
    ENTRY/DATA/MOTORの順番を維持して、AI投入用clean txtを作成。
    元HTMLは一切変更しない。
    """
    parts: List[str] = []

    for race_no in range(start_race, end_race + 1):
        parts.append(f"\n\n{'#'*90}")
        parts.append(f"# {race_no}R")
        parts.append(f"{'#'*90}")

        for page_name in ["entry", "data", "motor"]:
            html_path = race_dir / f"race_{race_no:02d}_{page_name}.html"
            txt_path = race_dir / f"race_{race_no:02d}_{page_name}.txt"

            parts.append(f"\n\n{'='*70}")
            parts.append(f"{race_no}R / {page_name.upper()}")
            parts.append(f"{'='*70}")

            if html_path.exists():
                html = html_path.read_text(encoding="utf-8", errors="ignore")
                parts.append(strip_html_to_clean_text(html))
            elif txt_path.exists():
                parts.append(txt_path.read_text(encoding="utf-8", errors="ignore"))
            else:
                parts.append("[MISSING]")

    clean_path = out_dir / f"all_races_{suffix}_clean.txt"
    clean_path.write_text("\n".join(parts), encoding="utf-8", errors="ignore")
    log[f"clean_{suffix}_txt"] = str(clean_path)



def make_clean_all(out_dir: Path, race_dir: Path, log: Dict):
    """
    AI投入用clean txtを1R〜12Rの1本にまとめる。
    元HTMLの分割mergedは残す。
    """
    parts: List[str] = []
    parts.append("# BOATERS AI予想用 CLEAN DATA")
    parts.append(f"stadium: {log.get('stadium_jp')} / {log.get('stadium_slug')}")
    parts.append(f"date: {log.get('date')}")
    parts.append("")
    parts.append("構成: 各Rごとに ENTRY → DATA → MOTOR")
    parts.append("ENTRY=出走表/節間成績, DATA=決まり手率/枠別/ST等, MOTOR=モーター情報/直近10走")
    parts.append("")

    for race_no in range(1, 13):
        parts.append(f"\n\n{'#'*100}")
        parts.append(f"# {race_no}R")
        parts.append(f"{'#'*100}")

        for page_name in ["entry", "data", "motor"]:
            parts.append(f"\n\n{'='*80}")
            parts.append(f"{race_no}R / {page_name.upper()}")
            parts.append(f"{'='*80}")

            html_path = race_dir / f"race_{race_no:02d}_{page_name}.html"
            txt_path = race_dir / f"race_{race_no:02d}_{page_name}.txt"

            if html_path.exists():
                html = html_path.read_text(encoding="utf-8", errors="ignore")
                cleaned = strip_html_to_clean_text(html)
                parts.append(cleaned if cleaned else "[EMPTY AFTER CLEAN]")
            elif txt_path.exists():
                parts.append(txt_path.read_text(encoding="utf-8", errors="ignore"))
            else:
                parts.append("[MISSING]")

    clean_all_txt = out_dir / "all_races_01_12_clean.txt"
    clean_all_txt.write_text("\n".join(parts), encoding="utf-8", errors="ignore")
    log["clean_01_12_txt"] = str(clean_all_txt)




SECTION_NAMES = ["ENTRY", "DATA", "MOTOR"]


def split_clean_races(text: str) -> Dict[int, str]:
    """
    all_races_01_12_clean.txt を # 1R〜# 12R で分割。
    """
    pattern = re.compile(r"(?m)^#\s*(\d{1,2})R\s*$")
    matches = list(pattern.finditer(text))
    races: Dict[int, str] = {}

    for i, m in enumerate(matches):
        race_no = int(m.group(1))
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        races[race_no] = text[start:end].strip()

    return races


def split_clean_sections(race_text: str) -> Dict[str, str]:
    """
    各R内を ENTRY / DATA / MOTOR に分割。
    """
    sections: Dict[str, str] = {k: "" for k in SECTION_NAMES}
    pattern = re.compile(r"(?m)^\d{1,2}R\s*/\s*(ENTRY|DATA|MOTOR)\s*$")
    matches = list(pattern.finditer(race_text))

    for i, m in enumerate(matches):
        sec = m.group(1)
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(race_text)
        block = race_text[start:end]
        block = re.sub(r"(?m)^=+\s*$", "", block)
        block = re.sub(r"(?m)^#+\s*$", "", block)
        sections[sec] = block.strip()

    return sections


def excel_clean_lines(block: str) -> List[str]:
    lines: List[str] = []
    blank = 0

    for raw in block.splitlines():
        line = raw.strip()
        if not line:
            blank += 1
            if blank <= 1:
                lines.append("")
            continue

        blank = 0
        line = re.sub(r"\s+", " ", line)
        lines.append(line)

    while lines and not lines[0]:
        lines.pop(0)
    while lines and not lines[-1]:
        lines.pop()

    return lines


def write_excel_section(ws, start_row: int, title: str, lines: List[str], fill_color: str) -> int:
    ws.cell(row=start_row, column=1, value=title)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)

    title_cell = ws.cell(row=start_row, column=1)
    title_cell.font = Font(bold=True, color="FFFFFF", size=12)
    title_cell.fill = PatternFill("solid", fgColor=fill_color)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    row = start_row + 1
    headers = ["No", "区分", "内容"]

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E5E7EB")
        c.alignment = Alignment(horizontal="center", vertical="center")

    row += 1

    for idx, line in enumerate(lines, start=1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value=line)
        row += 1

    return row + 2


def style_excel_sheet(ws):
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 120

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 18
        val = ws.cell(row=r, column=3).value
        if isinstance(val, str) and len(val) > 80:
            ws.row_dimensions[r].height = 34
        if isinstance(val, str) and len(val) > 180:
            ws.row_dimensions[r].height = 52


def create_excel_summary_sheet(wb: Workbook, races: Dict[int, str], log: Dict):
    ws = wb.create_sheet("Summary", 0)

    ws["A1"] = "BOATERS 取得データ Excel 変換結果"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws.merge_cells("A1:F1")

    ws["A2"] = f"{log.get('stadium_jp')} / {log.get('date')}"
    ws.merge_cells("A2:F2")

    headers = ["R", "シート", "ENTRY", "DATA", "MOTOR", "備考"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9EAF7")
        c.alignment = Alignment(horizontal="center")

    for row_idx, race_no in enumerate(range(1, 13), start=5):
        race_text = races.get(race_no, "")
        sections = split_clean_sections(race_text) if race_text else {k: "" for k in SECTION_NAMES}
        ws.cell(row=row_idx, column=1, value=f"{race_no}R")
        ws.cell(row=row_idx, column=2, value=f"{race_no}R")
        ws.cell(row=row_idx, column=3, value="OK" if sections.get("ENTRY") else "なし")
        ws.cell(row=row_idx, column=4, value="OK" if sections.get("DATA") else "なし")
        ws.cell(row=row_idx, column=5, value="OK" if sections.get("MOTOR") else "なし")
        ws.cell(row=row_idx, column=6, value="")

    for idx, width in enumerate([8, 14, 12, 12, 12, 30], start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=4, max_row=16, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def clean_txt_to_xlsx(clean_txt_path: Path, output_path: Path, log: Dict):
    """
    clean.txtから1R〜12RのExcelを作成。
    """
    if Workbook is None:
        print("[WARN] openpyxl が未インストールのためxlsx作成をスキップしました。")
        print("       py -m pip install openpyxl")
        log["xlsx_warning"] = "openpyxl not installed"
        return

    text = clean_txt_path.read_text(encoding="utf-8", errors="ignore")
    races = split_clean_races(text)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    create_excel_summary_sheet(wb, races, log)

    colors = {
        "ENTRY": "2563EB",
        "DATA": "16A34A",
        "MOTOR": "EA580C",
    }

    for race_no in range(1, 13):
        ws = wb.create_sheet(f"{race_no}R")

        ws["A1"] = f"{race_no}R"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="111827")
        ws.merge_cells("A1:C1")

        race_text = races.get(race_no, "")
        if not race_text:
            ws["A3"] = "このレースのデータが見つかりません"
            style_excel_sheet(ws)
            continue

        sections = split_clean_sections(race_text)

        row = 3
        for sec in SECTION_NAMES:
            lines = excel_clean_lines(sections.get(sec, ""))
            if not lines:
                lines = ["[データなし]"]
            row = write_excel_section(ws, row, sec, lines, colors[sec])

        style_excel_sheet(ws)

    wb.save(output_path)
    log["xlsx_01_12"] = str(output_path)



def append_file_text(parts: List[str], path: Path, header: str):
    if path.exists():
        parts.append(f"\n\n{'='*90}\n{header}\n{'='*90}\n")
        parts.append(path.read_text(encoding="utf-8", errors="ignore"))


def append_file_html(parts: List[str], path: Path, header: str):
    if path.exists():
        parts.append(f"<h1>{header}</h1>")
        parts.append(path.read_text(encoding="utf-8", errors="ignore"))
        parts.append("<hr>")


def combine_range(out_dir: Path, race_dir: Path, start_race: int, end_race: int, suffix: str, log: Dict):
    txt_parts: List[str] = []
    html_parts: List[str] = ["<html><head><meta charset='utf-8'></head><body>"]

    for race_no in range(start_race, end_race + 1):
        for page_name in ["entry", "data", "motor"]:
            txt_path = race_dir / f"race_{race_no:02d}_{page_name}.txt"
            html_path = race_dir / f"race_{race_no:02d}_{page_name}.html"
            header = f"{race_no}R / {page_name.upper()}"

            append_file_text(txt_parts, txt_path, header)
            append_file_html(html_parts, html_path, header)

    html_parts.append("</body></html>")

    merged_txt = out_dir / f"all_races_{suffix}_merged.txt"
    merged_html = out_dir / f"all_races_{suffix}_merged.html"

    merged_txt.write_text("\n".join(txt_parts), encoding="utf-8", errors="ignore")
    merged_html.write_text("\n".join(html_parts), encoding="utf-8", errors="ignore")

    log[f"merged_{suffix}_txt"] = str(merged_txt)
    log[f"merged_{suffix}_html"] = str(merged_html)


def make_report(out_dir: Path, log: Dict):
    lines: List[str] = []
    lines.append("# BOATERS 全12R 統合取得 v3.0 Report\n")
    lines.append(f"stadium_jp: {log.get('stadium_jp')}")
    lines.append(f"stadium_slug: {log.get('stadium_slug')}")
    lines.append(f"date: {log.get('date')}")
    lines.append(f"save_dir: {log.get('save_dir')}")
    lines.append("")
    lines.append(f"merged_01_06_html: {log.get('merged_01_06_html')}")
    lines.append(f"merged_01_06_txt: {log.get('merged_01_06_txt')}")
    lines.append(f"merged_07_12_html: {log.get('merged_07_12_html')}")
    lines.append(f"merged_07_12_txt: {log.get('merged_07_12_txt')}")
    lines.append(f"clean_01_12_txt: {log.get('clean_01_12_txt')}")
    lines.append(f"xlsx_01_12: {log.get('xlsx_01_12')}")
    lines.append("\n## Race Summary\n")
    lines.append("|R|entry|entry_core|data|motor|節間|entry hits|data hits|motor hits|entry current_url|")
    lines.append("|---:|---|---|---|---|---|---|---|---|---|")

    entry_map = {r["race_no"]: r for r in log.get("entry_results", [])}
    data_map = {r["race_no"]: r for r in log.get("data_results", [])}
    motor_map = {r["race_no"]: r for r in log.get("motor_results", [])}

    for race_no in range(1, 13):
        e = entry_map.get(race_no, {})
        d = data_map.get(race_no, {})
        m = motor_map.get(race_no, {})

        ea = e.get("analysis", {})
        da = d.get("analysis", {})
        ma = m.get("analysis", {})

        lines.append(
            f"|{race_no}R|{e.get('ok')}|{ea.get('has_entry_core')}|{d.get('ok')}|{m.get('ok')}|"
            f"{ea.get('setsukan_status')}|"
            f"{', '.join(ea.get('hits', [])[:10])}|"
            f"{', '.join(da.get('hits', [])[:8])}|"
            f"{', '.join(ma.get('hits', [])[:8])}|"
            f"{e.get('current_url')}|"
        )

    lines.append("\n## Notes\n")
    lines.append("- ENTRYは /data から「出走表」タブをクリックし、下までスクロールして本体テーブルを強制ロードします。")
    lines.append("- entry_core が False の場合は、まだ出走表本体ではありません。")
    lines.append("- 初日は節間成績が空で正常扱いです。")
    lines.append("- AI予想本番投入はノイズが気になる場合 all_races_01_12_clean.txt を推奨。")
    lines.append("- 元HTMLも保存されるため、cleanで不足があれば merged.html / races配下のHTMLへ戻れます。")

    report_path = out_dir / "summary_report.md"
    report_path.write_text("\n".join(lines), encoding="utf-8", errors="ignore")
    log["summary_report"] = str(report_path)


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--stadium", required=True, help="例: ashiya / 芦屋")
    parser.add_argument("--date", default=None, help="YYYY-MM-DD。省略時は日本時間の今日を自動使用")
    parser.add_argument("--root", default=str(BASE_SAVE_ROOT), help="保存ルート")
    parser.add_argument("--wait", type=float, default=6.0)
    parser.add_argument("--click-wait", type=float, default=1.5)
    parser.add_argument("--headed", action="store_true")
    parser.add_argument("--start-race", type=int, default=1)
    parser.add_argument("--end-race", type=int, default=12)
    args = parser.parse_args()

    stadium_jp = normalize_stadium(args.stadium)
    stadium_slug = stadium_to_slug(args.stadium)
    actual_date = args.date or today_jst()
    date_folder = date_to_folder(actual_date)

    root = Path(args.root)
    out_dir = root / stadium_jp / date_folder
    race_dir = out_dir / "races"

    out_dir.mkdir(parents=True, exist_ok=True)
    race_dir.mkdir(parents=True, exist_ok=True)

    wait_ms = int(args.wait * 1000)
    click_wait_ms = int(args.click_wait * 1000)

    log: Dict = {
        "version": "v3.0",
        "stadium_input": args.stadium,
        "stadium_jp": stadium_jp,
        "stadium_slug": stadium_slug,
        "date": actual_date,
        "save_dir": str(out_dir),
        "entry_results": [],
        "data_results": [],
        "motor_results": [],
    }

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=not args.headed)
        context = await browser.new_context(
            viewport={"width": 1365, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()

        for race_no in range(args.start_race, args.end_race + 1):
            print(f"\n========== {race_no}R ==========")

            entry_result = await capture_entry_by_click(page, stadium_slug, actual_date, race_no, race_dir, wait_ms)
            log["entry_results"].append(entry_result)

            data_result = await capture_data_page(page, stadium_slug, actual_date, race_no, race_dir, wait_ms, click_wait_ms)
            log["data_results"].append(data_result)

            motor_result = await capture_motor_page(page, stadium_slug, actual_date, race_no, race_dir, wait_ms)
            log["motor_results"].append(motor_result)

        await browser.close()

    combine_range(out_dir, race_dir, 1, 6, "01_06", log)
    combine_range(out_dir, race_dir, 7, 12, "07_12", log)

    # AI予想投入用clean txtは1R〜12Rを1本化
    make_clean_all(out_dir, race_dir, log)

    # Excelも同時作成
    clean_txt_path = out_dir / "all_races_01_12_clean.txt"
    xlsx_path = out_dir / "all_races_01_12_clean.xlsx"
    clean_txt_to_xlsx(clean_txt_path, xlsx_path, log)

    make_report(out_dir, log)

    (out_dir / "capture_log.json").write_text(
        json.dumps(log, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print("\n[DONE]")
    print(f"[SAVE_DIR] {out_dir}")
    print(f"[REPORT] {out_dir / 'summary_report.md'}")
    print(f"[MERGED 01-06 HTML] {out_dir / 'all_races_01_06_merged.html'}")
    print(f"[MERGED 07-12 HTML] {out_dir / 'all_races_07_12_merged.html'}")
    print(f"[CLEAN 01-12 TXT] {out_dir / 'all_races_01_12_clean.txt'}")
    print(f"[XLSX 01-12] {out_dir / 'all_races_01_12_clean.xlsx'}")


if __name__ == "__main__":
    asyncio.run(main())
